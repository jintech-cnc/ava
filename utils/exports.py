"""
Utils pour l'export des documents en PDF et Excel.

Tous les documents partagent désormais un design moderne avec :
  - En-tête entreprise (logo, nom, coordonnées, immatriculations)
  - Pied de page (mentions, page X/Y)
  - Palette de couleurs unifiée (#4f5eff = bleu Ava, #1c2130 = sombre, #f9fafc = gris clair)
  - Devise configurable par entreprise
"""
import io
import os
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image,
    PageBreak, KeepTogether,
)
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.pdfgen import canvas
from reportlab.platypus.flowables import HRFlowable
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from django.utils import timezone


# ==========================================
# PALETTE & CONSTANTES
# ==========================================

PRIMARY = '#4f5eff'
PRIMARY_DARK = '#3741d4'
DARK = '#1c2130'
LIGHT_BG = '#f9fafc'
GREY_BORDER = '#e5e8ef'
GREY_TEXT = '#6b7280'
DANGER = '#dc2626'
WARNING = '#f59e0b'
SUCCESS = '#10b981'

DEFAULT_COMPANY = {
    'nom': 'Ava',
    'slogan': '',
    'logo': None,
    'adresse': '',
    'telephone': '',
    'email': '',
    'site_web': '',
    'rccm': '',
    'id_national': '',
    'numero_impot': '',
    'devise': 'FC',
    'separateur_milliers': ' ',
    'prefixe_symbole': True,
    'mentions_footer': '',
}


# ==========================================
# HELPERS COMPANY
# ==========================================

def _company_or_default(company):
    """Renvoie un dict company (utilise les valeurs par défaut si None)."""
    if company is None:
        return dict(DEFAULT_COMPANY)
    if isinstance(company, dict):
        # Fusionne avec les valeurs par défaut
        return {**DEFAULT_COMPANY, **company}
    return {
        'nom': company.nom,
        'slogan': company.slogan,
        'logo': company.logo,
        'adresse': company.adresse,
        'telephone': company.telephone,
        'email': company.email,
        'site_web': company.site_web,
        'rccm': company.rccm,
        'id_national': company.id_national,
        'numero_impot': company.numero_impot,
        'devise': company.devise,
        'separateur_milliers': company.separateur_milliers,
        'prefixe_symbole': company.prefixe_symbole,
        'mentions_footer': company.mentions_footer,
    }


def format_montant(value, company):
    """Formate un montant avec la devise de l'entreprise."""
    info = _company_or_default(company)
    try:
        v = float(value)
        formatted = f"{v:,.2f}".replace(',', 'X').replace('.', ',').replace('X', info['separateur_milliers'])
    except (ValueError, TypeError):
        formatted = str(value)
    if info['prefixe_symbole']:
        return f"{info['devise']} {formatted}"
    return f"{formatted} {info['devise']}"


# ==========================================
# PDF : header/footer + canvas
# ==========================================

class BrandedDocTemplate(SimpleDocTemplate):
    """Document template avec en-tête et pied de page entreprise."""
    def __init__(self, *args, company=None, document_title='', document_ref='', **kwargs):
        super().__init__(*args, **kwargs)
        self.company = _company_or_default(company)
        self.document_title = document_title
        self.document_ref = document_ref

    def afterFlowable(self, flowable):
        pass

    def _draw_header(self, canv, doc):
        canv.saveState()
        company = self.company
        w, h = A4
        margin = 15 * mm
        y = h - 12 * mm

        # Bloc logo (gauche)
        x_logo = margin
        logo_w = 30 * mm
        logo_path = None
        if company.get('logo') and hasattr(company['logo'], 'path'):
            logo_path = company['logo'].path
        elif isinstance(company.get('logo'), str) and os.path.isfile(company['logo']):
            logo_path = company['logo']

        if logo_path and os.path.isfile(logo_path):
            try:
                canv.drawImage(
                    logo_path, x_logo, y - 22 * mm,
                    width=logo_w, height=22 * mm,
                    preserveAspectRatio=True, mask='auto',
                )
                x_text = x_logo + logo_w + 5 * mm
            except Exception:
                x_text = x_logo
        else:
            # Pas de logo : carré coloré avec initiale
            canv.setFillColor(colors.HexColor(PRIMARY))
            canv.roundRect(x_logo, y - 22 * mm, 22 * mm, 22 * mm, 2 * mm, fill=1, stroke=0)
            canv.setFillColor(colors.white)
            canv.setFont('Helvetica-Bold', 18)
            initial = company['nom'][:1].upper() if company['nom'] else 'A'
            canv.drawCentredString(x_logo + 11 * mm, y - 14 * mm, initial)
            x_text = x_logo + 25 * mm

        # Bloc nom entreprise (gauche)
        canv.setFillColor(colors.HexColor(DARK))
        canv.setFont('Helvetica-Bold', 14)
        canv.drawString(x_text, y - 6 * mm, company['nom'])
        if company.get('slogan'):
            canv.setFont('Helvetica-Oblique', 8)
            canv.setFillColor(colors.HexColor(GREY_TEXT))
            canv.drawString(x_text, y - 11 * mm, company['slogan'])

        # Coordonnées (gauche, en petit)
        canv.setFont('Helvetica', 7.5)
        canv.setFillColor(colors.HexColor(GREY_TEXT))
        line_y = y - 16 * mm
        if company.get('adresse'):
            # Découpe l'adresse en lignes (max 60 chars)
            addr = company['adresse'].replace('\n', ' / ')
            for chunk in [addr[i:i+55] for i in range(0, min(len(addr), 110), 55)]:
                canv.drawString(x_text, line_y, chunk)
                line_y -= 3.5 * mm
        if company.get('telephone') or company.get('email'):
            line = []
            if company.get('telephone'):
                line.append(f"Tél: {company['telephone']}")
            if company.get('email'):
                line.append(company['email'])
            canv.drawString(x_text, line_y, '  •  '.join(line))
            line_y -= 3.5 * mm
        if company.get('site_web'):
            canv.drawString(x_text, line_y, company['site_web'])

        # Bloc titre document (droite)
        right_x = w - margin
        canv.setFillColor(colors.HexColor(PRIMARY))
        canv.setFont('Helvetica-Bold', 16)
        title = self.document_title or 'DOCUMENT'
        canv.drawRightString(right_x, y - 6 * mm, title)

        if self.document_ref:
            canv.setFont('Helvetica', 9)
            canv.setFillColor(colors.HexColor(DARK))
            canv.drawRightString(right_x, y - 12 * mm, f"Réf: {self.document_ref}")

        canv.setFont('Helvetica', 8)
        canv.setFillColor(colors.HexColor(GREY_TEXT))
        date_str = timezone.now().strftime('%d/%m/%Y à %H:%M')
        canv.drawRightString(right_x, y - 17 * mm, f"Émis le {date_str}")

        # Filet horizontal décoratif (barre dégradée : primaire + clair)
        canv.setStrokeColor(colors.HexColor(PRIMARY))
        canv.setLineWidth(1.5)
        canv.line(margin, y - 25 * mm, right_x, y - 25 * mm)
        canv.setStrokeColor(colors.HexColor(PRIMARY))
        canv.setLineWidth(0.5)
        canv.line(margin, y - 26 * mm, margin + 50 * mm, y - 26 * mm)

        # Immatriculations sous le filet (gauche)
        imm_y = y - 32 * mm
        canv.setFont('Helvetica', 7)
        canv.setFillColor(colors.HexColor(GREY_TEXT))
        imm_parts = []
        if company.get('rccm'):
            imm_parts.append(f"RCCM: {company['rccm']}")
        if company.get('id_national'):
            imm_parts.append(f"ID Nat: {company['id_national']}")
        if company.get('numero_impot'):
            imm_parts.append(f"N° Impôt: {company['numero_impot']}")
        if imm_parts:
            canv.drawString(margin, imm_y, '   |   '.join(imm_parts))

        canv.restoreState()

    def _draw_footer(self, canv, doc):
        canv.saveState()
        company = self.company
        w, h = A4
        margin = 15 * mm
        y = 12 * mm

        # Filet
        canv.setStrokeColor(colors.HexColor(GREY_BORDER))
        canv.setLineWidth(0.5)
        canv.line(margin, y + 8 * mm, w - margin, y + 8 * mm)

        # Mentions gauche
        canv.setFont('Helvetica', 7)
        canv.setFillColor(colors.HexColor(GREY_TEXT))
        mentions = company.get('mentions_footer') or f"{company['nom']} — Document généré par Ava"
        # Tronque si trop long
        if len(mentions) > 90:
            mentions = mentions[:87] + '...'
        canv.drawString(margin, y + 3 * mm, mentions)

        # Pagination droite
        canv.setFont('Helvetica-Bold', 7)
        canv.setFillColor(colors.HexColor(PRIMARY))
        canv.drawRightString(w - margin, y + 3 * mm, f"Page {doc.page} / {{nbpages}}")

        canv.restoreState()

    def handle_pageBegin(self):
        # Dessine header + footer sur chaque nouvelle page
        self._draw_header(self.canv, self)
        self._draw_footer(self.canv, self)
        # Continue le comportement normal de SimpleDocTemplate
        from reportlab.platypus.doctemplate import SimpleDocTemplate
        SimpleDocTemplate.handle_pageBegin(self)


def _build_doc(buffer, company, title='', ref=''):
    """Helper : crée un BrandedDocTemplate configuré."""
    doc = BrandedDocTemplate(
        buffer, pagesize=A4,
        rightMargin=15*mm, leftMargin=15*mm,
        topMargin=45*mm, bottomMargin=18*mm,
        company=company, document_title=title, document_ref=ref,
    )
    return doc


# ==========================================
# PDF STYLES
# ==========================================

def _get_base_style():
    styles = getSampleStyleSheet()
    if 'AvaSection' not in styles:
        styles.add(ParagraphStyle(
            name='AvaSection',
            parent=styles['Heading2'],
            fontSize=12, leading=14,
            textColor=colors.HexColor(DARK),
            spaceBefore=10, spaceAfter=6,
        ))
    if 'AvaSubtitle' not in styles:
        styles.add(ParagraphStyle(
            name='AvaSubtitle',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor(GREY_TEXT),
        ))
    if 'AvaInfo' not in styles:
        styles.add(ParagraphStyle(
            name='AvaInfo',
            parent=styles['Normal'],
            fontSize=9, leading=12,
            textColor=colors.HexColor(DARK),
        ))
    if 'AvaAlert' not in styles:
        styles.add(ParagraphStyle(
            name='AvaAlert',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor(DANGER),
        ))
    return styles


def _table_style_base():
    """Style de tableau unifié : header primaire, lignes alternées, bordures fines."""
    return TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(PRIMARY)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        # Corps
        ('FONTSIZE', (0, 1), (-1, -1), 8.5),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
        # Bordures
        ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor(PRIMARY_DARK)),
        ('GRID', (0, 1), (-1, -1), 0.3, colors.HexColor(GREY_BORDER)),
        # Alternance
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor(LIGHT_BG)]),
        # VALIGN
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])


# ==========================================
# PDF EXPORTS
# ==========================================

def generate_product_list_pdf(products, company=None, language='fr'):
    """Liste des produits."""
    styles = _get_base_style()
    buffer = io.BytesIO()
    doc = _build_doc(buffer, company, title='LISTE DES PRODUITS')

    title_map = {
        'fr': 'Catalogue complet',
        'en': 'Full catalog',
        'zh-hans': '完整目录',
        'hi': 'पूरी सूची',
    }

    elements = [
        Paragraph(title_map.get(language, title_map['fr']), styles['AvaSubtitle']),
        Spacer(1, 6*mm),
    ]

    data = [['Réf.', 'Désignation', 'Catégorie', 'Unité', 'Prix unitaire', 'Stock', 'État']]
    for p in products:
        cat = p.categorie.nom if p.categorie else '—'
        etat = 'Actif' if p.actif else 'Inactif'
        try:
            stock = str(int(p.quantite_totale))
        except Exception:
            stock = str(p.stock_total) if hasattr(p, 'stock_total') else '0'
        data.append([
            p.reference,
            p.nom[:45],
            cat[:18],
            p.unite,
            format_montant(p.prix_unitaire, company),
            stock,
            etat,
        ])

    table = Table(data, colWidths=[25*mm, 60*mm, 28*mm, 18*mm, 30*mm, 15*mm, 14*mm], repeatRows=1)
    table.setStyle(_table_style_base())
    table.setStyle(TableStyle([
        ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
        ('ALIGN', (5, 1), (5, -1), 'CENTER'),
        ('ALIGN', (6, 1), (6, -1), 'CENTER'),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_order_pdf(order, company=None, language='fr'):
    """Facture / commande client."""
    styles = _get_base_style()
    buffer = io.BytesIO()
    doc = _build_doc(buffer, company, title='FACTURE', ref=f"#{order.pk}")

    elements = []

    # Bloc client (encadré)
    client_info = [
        Paragraph(f"<b>DESTINATAIRE</b>", styles['AvaSubtitle']),
        Paragraph(f"<b>{order.client.nom}</b>", styles['AvaInfo']),
    ]
    if hasattr(order.client, 'telephone') and order.client.telephone:
        client_info.append(Paragraph(f"Tél : {order.client.telephone}", styles['AvaInfo']))
    if hasattr(order.client, 'email') and order.client.email:
        client_info.append(Paragraph(f"Email : {order.client.email}", styles['AvaInfo']))
    if hasattr(order.client, 'adresse') and order.client.adresse:
        client_info.append(Paragraph(f"{order.client.adresse}", styles['AvaInfo']))

    # Bloc émission
    emission_info = [
        Paragraph(f"<b>INFORMATIONS</b>", styles['AvaSubtitle']),
        Paragraph(f"Date d'émission : <b>{order.cree_le.strftime('%d/%m/%Y')}</b>", styles['AvaInfo']),
        Paragraph(f"Statut : <b>{order.get_statut_display()}</b>", styles['AvaInfo']),
    ]
    if hasattr(order, 'numero') and order.numero:
        emission_info.insert(1, Paragraph(f"Numéro : <b>{order.numero}</b>", styles['AvaInfo']))

    bloc = Table([[client_info, emission_info]], colWidths=[95*mm, 70*mm])
    bloc.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(LIGHT_BG)),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor(GREY_BORDER)),
        ('INNERGRID', (0, 0), (-1, -1), 0.3, colors.HexColor(GREY_BORDER)),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(bloc)
    elements.append(Spacer(1, 8*mm))

    # Lignes de facture
    data = [['#', 'Désignation', 'Qté', 'Prix unit.', 'Total']]
    total_qte = 0
    for idx, ligne in enumerate(order.lignes.all(), 1):
        data.append([
            str(idx),
            ligne.produit.nom,
            str(ligne.quantite),
            format_montant(ligne.prix_unitaire, company),
            format_montant(ligne.sous_total, company),
        ])
        total_qte += ligne.quantite

    # Totaux
    data.append(['', '', '', Paragraph('<b>Sous-total HT</b>', styles['AvaInfo']),
                 Paragraph(f"<b>{format_montant(order.montant_total, company)}</b>", styles['AvaInfo'])])
    data.append(['', '', '', Paragraph('<b>TOTAL TTC</b>', styles['AvaInfo']),
                 Paragraph(f"<b>{format_montant(order.montant_total, company)}</b>", styles['AvaInfo'])])

    table = Table(data, colWidths=[12*mm, 75*mm, 18*mm, 35*mm, 35*mm], repeatRows=1)
    style = _table_style_base()
    style.add('ALIGN', (2, 1), (4, -1), 'RIGHT')
    style.add('ALIGN', (2, 0), (2, 0), 'CENTER')
    style.add('LINEABOVE', (3, -2), (4, -1), 1, colors.HexColor(PRIMARY))
    style.add('BACKGROUND', (3, -2), (-1, -1), colors.HexColor(LIGHT_BG))
    style.add('FONTNAME', (3, -2), (-1, -1), 'Helvetica-Bold')
    style.add('TEXTCOLOR', (3, -2), (-1, -1), colors.HexColor(DARK))
    table.setStyle(style)
    elements.append(table)

    if hasattr(order, 'notes') and order.notes:
        elements.append(Spacer(1, 6*mm))
        elements.append(Paragraph('<b>Notes :</b>', styles['AvaSection']))
        elements.append(Paragraph(order.notes, styles['AvaInfo']))

    # Zone signature
    elements.append(Spacer(1, 20*mm))
    sig = Table([
        [Paragraph('<b>Signature client</b>', styles['AvaSubtitle']),
         Paragraph('<b>Cachet & signature</b>', styles['AvaSubtitle'])],
        ['', ''],
        ['', ''],
    ], colWidths=[80*mm, 80*mm], rowHeights=[5*mm, 18*mm, 5*mm])
    sig.setStyle(TableStyle([
        ('LINEBELOW', (0, 1), (0, 1), 0.5, colors.HexColor(DARK)),
        ('LINEBELOW', (1, 1), (1, 1), 0.5, colors.HexColor(DARK)),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
    ]))
    elements.append(sig)

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_purchase_order_pdf(po, company=None, language='fr'):
    """Bon de commande fournisseur."""
    styles = _get_base_style()
    buffer = io.BytesIO()
    doc = _build_doc(buffer, company, title='BON DE COMMANDE', ref=po.reference)

    elements = []

    # Bloc fournisseur
    fournisseur_info = [
        Paragraph('<b>FOURNISSEUR</b>', styles['AvaSubtitle']),
        Paragraph(f"<b>{po.fournisseur.nom}</b>", styles['AvaInfo']),
    ]
    if po.fournisseur.contact:
        fournisseur_info.append(Paragraph(f"Contact : {po.fournisseur.contact}", styles['AvaInfo']))
    if po.fournisseur.telephone:
        fournisseur_info.append(Paragraph(f"Tél : {po.fournisseur.telephone}", styles['AvaInfo']))
    if po.fournisseur.email:
        fournisseur_info.append(Paragraph(f"Email : {po.fournisseur.email}", styles['AvaInfo']))
    if po.fournisseur.adresse:
        fournisseur_info.append(Paragraph(po.fournisseur.adresse, styles['AvaInfo']))

    # Bloc interne
    interne_info = [
        Paragraph('<b>INFORMATIONS</b>', styles['AvaSubtitle']),
        Paragraph(f"Date : <b>{po.cree_le.strftime('%d/%m/%Y')}</b>", styles['AvaInfo']),
        Paragraph(f"Statut : <b>{po.get_statut_display()}</b>", styles['AvaInfo']),
    ]
    if po.entrepot:
        interne_info.append(Paragraph(f"Entrepôt : <b>{po.entrepot.nom}</b>", styles['AvaInfo']))
    if po.date_livraison_prevue:
        interne_info.append(Paragraph(f"Livraison prévue : <b>{po.date_livraison_prevue.strftime('%d/%m/%Y')}</b>", styles['AvaInfo']))

    bloc = Table([[fournisseur_info, interne_info]], colWidths=[95*mm, 70*mm])
    bloc.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(LIGHT_BG)),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor(GREY_BORDER)),
        ('INNERGRID', (0, 0), (-1, -1), 0.3, colors.HexColor(GREY_BORDER)),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(bloc)
    elements.append(Spacer(1, 8*mm))

    data = [['#', 'Référence', 'Désignation', 'Qté cde', 'Qté reçue', 'Prix unit.', 'Total']]
    total = 0
    for idx, ligne in enumerate(po.lignes.all(), 1):
        data.append([
            str(idx),
            ligne.produit.reference,
            ligne.produit.nom,
            str(ligne.quantite_commandee),
            str(ligne.quantite_recue),
            format_montant(ligne.prix_unitaire, company),
            format_montant(ligne.montant_ligne, company),
        ])
        total += float(ligne.montant_ligne or 0)

    data.append(['', '', '', '', '', Paragraph('<b>TOTAL</b>', styles['AvaInfo']),
                 Paragraph(f"<b>{format_montant(total, company)}</b>", styles['AvaInfo'])])

    table = Table(data, colWidths=[10*mm, 22*mm, 50*mm, 18*mm, 18*mm, 28*mm, 30*mm], repeatRows=1)
    style = _table_style_base()
    style.add('ALIGN', (3, 1), (6, -1), 'RIGHT')
    style.add('ALIGN', (3, 0), (4, 0), 'CENTER')
    style.add('LINEABOVE', (5, -1), (6, -1), 1, colors.HexColor(PRIMARY))
    style.add('BACKGROUND', (5, -1), (-1, -1), colors.HexColor(LIGHT_BG))
    table.setStyle(style)
    elements.append(table)

    if po.notes:
        elements.append(Spacer(1, 6*mm))
        elements.append(Paragraph('<b>Notes :</b>', styles['AvaSection']))
        elements.append(Paragraph(po.notes, styles['AvaInfo']))

    elements.append(Spacer(1, 15*mm))
    sig = Table([
        [Paragraph('<b>Préparé par</b>', styles['AvaSubtitle']),
         Paragraph('<b>Approuvé par</b>', styles['AvaSubtitle']),
         Paragraph('<b>Réception fournisseur</b>', styles['AvaSubtitle'])],
        ['', '', ''],
        ['', '', ''],
    ], colWidths=[55*mm, 55*mm, 55*mm], rowHeights=[5*mm, 18*mm, 5*mm])
    sig.setStyle(TableStyle([
        ('LINEBELOW', (0, 1), (-1, 1), 0.5, colors.HexColor(DARK)),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
    ]))
    elements.append(sig)

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_movement_pdf(movements, company=None, language='fr'):
    """Historique des mouvements de stock."""
    styles = _get_base_style()
    buffer = io.BytesIO()
    doc = _build_doc(buffer, company, title='MOUVEMENTS DE STOCK')

    elements = [
        Paragraph('Historique complet des mouvements', styles['AvaSubtitle']),
        Spacer(1, 6*mm),
    ]

    data = [['Date', 'Produit', 'Entrepôt', 'Type', 'Qté', 'Effectué par', 'Motif']]
    for m in movements:
        type_display = {
            'entree': 'Entrée', 'sortie': 'Sortie',
            'ajustement': 'Ajustement', 'transfert': 'Transfert',
        }.get(m.type_mouvement, m.type_mouvement)
        qte_display = f"+{m.quantite}" if m.type_mouvement == 'entree' else f"−{abs(m.quantite)}"
        data.append([
            m.cree_le.strftime('%d/%m/%Y %H:%M'),
            m.produit.nom[:35],
            str(m.entrepot)[:20],
            type_display,
            qte_display,
            str(m.effectue_par)[:18] if m.effectue_par else '—',
            (m.motif or '—')[:30],
        ])

    table = Table(data, colWidths=[28*mm, 42*mm, 25*mm, 22*mm, 18*mm, 25*mm, 25*mm], repeatRows=1)
    table.setStyle(_table_style_base())
    table.setStyle(TableStyle([
        ('ALIGN', (4, 1), (4, -1), 'CENTER'),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_requisition_pdf(lignes, company=None, language='fr'):
    """Fiche de réquisition provisoire pré-remplie."""
    styles = _get_base_style()
    buffer = io.BytesIO()
    doc = _build_doc(buffer, company, title='FICHE DE RÉQUISITION')

    elements = []
    elements.append(Paragraph(
        '<b>Propositions de réapprovisionnement</b> — produits sous le seuil minimum',
        styles['AvaSubtitle']
    ))
    elements.append(Spacer(1, 6*mm))

    if not lignes:
        elements.append(Paragraph(
            '✓ Aucun produit en alerte. Le stock est suffisant.',
            styles['AvaInfo']
        ))
    else:
        # En-tête encadré avec alerte
        total = sum(l['estimation'] for l in lignes)
        info = Table([[
            Paragraph(f"<b>{len(lignes)}</b> produit(s) en alerte", styles['AvaInfo']),
            Paragraph(f"Estimation totale : <b>{format_montant(total, company)}</b>", styles['AvaInfo']),
        ]], colWidths=[90*mm, 80*mm])
        info.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(WARNING)),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(info)
        elements.append(Spacer(1, 4*mm))

        data = [['#', 'Référence', 'Produit', 'Catégorie', 'Stock', 'Seuil', 'Qté cde', 'Fournisseur', 'Estim.']]

        for idx, l in enumerate(lignes, 1):
            fournisseur = l['fournisseur'].nom if l['fournisseur'] else '—'
            data.append([
                str(idx),
                l['produit'].reference,
                l['produit'].nom[:35],
                (l['produit'].categorie.nom if l['produit'].categorie else '—')[:15],
                str(l['stock_actuel']),
                str(l['produit'].seuil_alerte),
                str(l['qte_a_commander']),
                fournisseur[:18],
                format_montant(l['estimation'], company),
            ])

        data.append([
            '', '', '', '', '', '', '',
            Paragraph('<b>TOTAL ESTIMÉ</b>', styles['AvaInfo']),
            Paragraph(f"<b>{format_montant(total, company)}</b>", styles['AvaInfo']),
        ])

        table = Table(data, colWidths=[8*mm, 22*mm, 38*mm, 22*mm, 14*mm, 14*mm, 16*mm, 25*mm, 22*mm], repeatRows=1)
        style = _table_style_base()
        style.add('ALIGN', (4, 1), (6, -1), 'CENTER')
        style.add('ALIGN', (8, 1), (8, -1), 'RIGHT')
        style.add('LINEABOVE', (7, -1), (8, -1), 1, colors.HexColor(PRIMARY))
        style.add('BACKGROUND', (7, -1), (-1, -1), colors.HexColor(LIGHT_BG))
        table.setStyle(style)
        elements.append(table)

        # Zone d'approbation
        elements.append(Spacer(1, 15*mm))
        sig = Table([
            [Paragraph('<b>Demandé par</b>', styles['AvaSubtitle']),
             Paragraph('<b>Approuvé par</b>', styles['AvaSubtitle']),
             Paragraph('<b>Direction</b>', styles['AvaSubtitle'])],
            ['', '', ''],
            ['', '', ''],
        ], colWidths=[55*mm, 55*mm, 55*mm], rowHeights=[5*mm, 18*mm, 5*mm])
        sig.setStyle(TableStyle([
            ('LINEBELOW', (0, 1), (-1, 1), 0.5, colors.HexColor(DARK)),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ]))
        elements.append(sig)

    doc.build(elements)
    buffer.seek(0)
    return buffer


# ==========================================
# EXCEL EXPORTS
# ==========================================

def _xl_company_header(ws, company, title, ref=''):
    """Écrit l'en-tête entreprise dans une feuille Excel."""
    company = _company_or_default(company)
    row = 1

    # Nom entreprise (gros, primary)
    cell = ws.cell(row=row, column=1, value=company['nom'])
    cell.font = Font(name='Calibri', size=18, bold=True, color=PRIMARY.lstrip('#').upper())
    cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1

    if company.get('slogan'):
        ws.cell(row=row, column=1, value=company['slogan']).font = Font(italic=True, color=GREY_TEXT.lstrip('#').upper())
        row += 1

    if company.get('adresse'):
        ws.cell(row=row, column=1, value=company['adresse']).font = Font(size=9, color=GREY_TEXT.lstrip('#').upper())
        row += 1

    coord_parts = []
    if company.get('telephone'):
        coord_parts.append(f"Tél: {company['telephone']}")
    if company.get('email'):
        coord_parts.append(company['email'])
    if company.get('site_web'):
        coord_parts.append(company['site_web'])
    if coord_parts:
        ws.cell(row=row, column=1, value='  •  '.join(coord_parts)).font = Font(size=9, color=GREY_TEXT.lstrip('#').upper())
        row += 1

    imm_parts = []
    if company.get('rccm'):
        imm_parts.append(f"RCCM: {company['rccm']}")
    if company.get('id_national'):
        imm_parts.append(f"ID Nat: {company['id_national']}")
    if company.get('numero_impot'):
        imm_parts.append(f"N° Impôt: {company['numero_impot']}")
    if imm_parts:
        ws.cell(row=row, column=1, value='   |   '.join(imm_parts)).font = Font(size=8, color=GREY_TEXT.lstrip('#').upper())
        row += 1

    row += 1  # espace

    # Titre du document (en gros, primary)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(name='Calibri', size=14, bold=True, color=DARK.lstrip('#').upper())
    if ref:
        cell2 = ws.cell(row=row, column=2, value=ref)
        cell2.font = Font(name='Calibri', size=12, bold=True, color=PRIMARY.lstrip('#').upper())
        cell2.alignment = Alignment(horizontal='right')
    row += 1

    # Date d'émission
    cell = ws.cell(row=row, column=1, value=f"Émis le {timezone.now().strftime('%d/%m/%Y à %H:%M')}")
    cell.font = Font(size=9, italic=True, color=GREY_TEXT.lstrip('#').upper())
    row += 2

    return row


def _xl_border():
    thin = Side(style='thin', color=GREY_BORDER.lstrip('#').upper())
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _xl_header_fill():
    return PatternFill(start_color=PRIMARY.lstrip('#').upper(),
                       end_color=PRIMARY.lstrip('#').upper(), fill_type='solid')


def _xl_alt_fill():
    return PatternFill(start_color=LIGHT_BG.lstrip('#').upper(),
                       end_color=LIGHT_BG.lstrip('#').upper(), fill_type='solid')


def _xl_total_fill():
    return PatternFill(start_color=PRIMARY.lstrip('#').upper(),
                       end_color=PRIMARY.lstrip('#').upper(), fill_type='solid')


def _style_xl_header(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.fill = _xl_header_fill()
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _xl_border()


def _style_xl_row(ws, row, ncols, alt=False):
    fill = _xl_alt_fill() if alt else None
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(size=9)
        cell.border = _xl_border()
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if fill:
            cell.fill = fill


def generate_product_list_excel(products, company=None, language='fr'):
    """Liste des produits en Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Produits'

    start_row = _xl_company_header(ws, company, 'LISTE DES PRODUITS')

    headers = ['Référence', 'Désignation', 'Catégorie', 'Unité', 'Prix unitaire',
               'Stock total', 'Seuil alerte', 'État']
    for col, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=col, value=h)
    _style_xl_header(ws, start_row, len(headers))

    info = _company_or_default(company)
    for i, p in enumerate(products):
        row = start_row + 1 + i
        try:
            stock = int(p.quantite_totale)
        except Exception:
            stock = int(getattr(p, 'stock_total', 0) or 0)
        data = [
            p.reference,
            p.nom,
            p.categorie.nom if p.categorie else '',
            p.unite,
            format_montant(p.prix_unitaire, company),
            stock,
            p.seuil_alerte,
            'Actif' if p.actif else 'Inactif',
        ]
        for col, v in enumerate(data, 1):
            ws.cell(row=row, column=col, value=v)
        _style_xl_row(ws, row, len(headers), alt=(i % 2 == 1))

        if col == 5:
            ws.cell(row=row, column=col).number_format = f'"{info["devise"]}" # ##0.00'
        if col in (6, 7):
            ws.cell(row=row, column=col).number_format = '# ##0'

    # Largeurs
    widths = [15, 35, 18, 10, 18, 12, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Figer les volets
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_order_excel(order, company=None, language='fr'):
    """Commande client en Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Commande_{order.pk}"

    ref = order.numero if hasattr(order, 'numero') and order.numero else f"#{order.pk}"
    start_row = _xl_company_header(ws, company, 'FACTURE', ref=ref)

    # Bloc client
    ws.cell(row=start_row, column=1, value='DESTINATAIRE').font = Font(bold=True, size=10, color=PRIMARY.lstrip('#').upper())
    start_row += 1
    ws.cell(row=start_row, column=1, value=order.client.nom).font = Font(bold=True, size=10)
    start_row += 1
    if hasattr(order.client, 'telephone') and order.client.telephone:
        ws.cell(row=start_row, column=1, value=f"Tél : {order.client.telephone}")
        start_row += 1
    if hasattr(order.client, 'email') and order.client.email:
        ws.cell(row=start_row, column=1, value=f"Email : {order.client.email}")
        start_row += 1
    ws.cell(row=start_row, column=1, value=f"Date : {order.cree_le.strftime('%d/%m/%Y')}")
    start_row += 2

    headers = ['#', 'Désignation', 'Qté', 'Prix unitaire', 'Sous-total']
    for col, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=col, value=h)
    _style_xl_header(ws, start_row, len(headers))

    info = _company_or_default(company)
    for i, ligne in enumerate(order.lignes.all()):
        row = start_row + 1 + i
        data = [
            i + 1,
            ligne.produit.nom,
            ligne.quantite,
            format_montant(ligne.prix_unitaire, company),
            format_montant(ligne.sous_total, company),
        ]
        for col, v in enumerate(data, 1):
            ws.cell(row=row, column=col, value=v)
        _style_xl_row(ws, row, len(headers), alt=(i % 2 == 1))

    total_row = start_row + 1 + order.lignes.count()
    ws.cell(row=total_row, column=4, value='TOTAL TTC').font = Font(bold=True, color='FFFFFF')
    ws.cell(row=total_row, column=4).fill = _xl_total_fill()
    ws.cell(row=total_row, column=4).alignment = Alignment(horizontal='right')
    ws.cell(row=total_row, column=4).border = _xl_border()
    ws.cell(row=total_row, column=5, value=format_montant(order.montant_total, company))
    ws.cell(row=total_row, column=5).font = Font(bold=True, color='FFFFFF')
    ws.cell(row=total_row, column=5).fill = _xl_total_fill()
    ws.cell(row=total_row, column=5).border = _xl_border()

    widths = [6, 45, 10, 18, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_purchase_order_excel(po, company=None, language='fr'):
    """Bon de commande fournisseur en Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"BC_{po.reference}"

    start_row = _xl_company_header(ws, company, 'BON DE COMMANDE', ref=po.reference)

    ws.cell(row=start_row, column=1, value=f"Fournisseur : {po.fournisseur.nom}").font = Font(bold=True)
    start_row += 1
    ws.cell(row=start_row, column=1, value=f"Date : {po.cree_le.strftime('%d/%m/%Y')}")
    start_row += 1
    ws.cell(row=start_row, column=1, value=f"Statut : {po.get_statut_display()}")
    start_row += 2

    headers = ['#', 'Référence', 'Désignation', 'Qté commandée', 'Qté reçue', 'Prix unitaire', 'Total']
    for col, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=col, value=h)
    _style_xl_header(ws, start_row, len(headers))

    total = 0
    for i, ligne in enumerate(po.lignes.all()):
        row = start_row + 1 + i
        data = [
            i + 1,
            ligne.produit.reference,
            ligne.produit.nom,
            ligne.quantite_commandee,
            ligne.quantite_recue,
            format_montant(ligne.prix_unitaire, company),
            format_montant(ligne.montant_ligne, company),
        ]
        for col, v in enumerate(data, 1):
            ws.cell(row=row, column=col, value=v)
        _style_xl_row(ws, row, len(headers), alt=(i % 2 == 1))
        total += float(ligne.montant_ligne or 0)

    total_row = start_row + 1 + po.lignes.count()
    ws.cell(row=total_row, column=6, value='TOTAL').font = Font(bold=True, color='FFFFFF')
    ws.cell(row=total_row, column=6).fill = _xl_total_fill()
    ws.cell(row=total_row, column=6).alignment = Alignment(horizontal='right')
    ws.cell(row=total_row, column=6).border = _xl_border()
    ws.cell(row=total_row, column=7, value=format_montant(total, company))
    ws.cell(row=total_row, column=7).font = Font(bold=True, color='FFFFFF')
    ws.cell(row=total_row, column=7).fill = _xl_total_fill()
    ws.cell(row=total_row, column=7).border = _xl_border()

    widths = [6, 18, 40, 16, 14, 18, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_movement_excel(movements, company=None, language='fr'):
    """Mouvements de stock en Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Mouvements'

    start_row = _xl_company_header(ws, company, 'MOUVEMENTS DE STOCK')

    headers = ['Date', 'Produit', 'Référence', 'Entrepôt', 'Type', 'Quantité', 'Effectué par', 'Motif']
    for col, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=col, value=h)
    _style_xl_header(ws, start_row, len(headers))

    type_map = {'entree': 'Entrée', 'sortie': 'Sortie', 'ajustement': 'Ajustement', 'transfert': 'Transfert'}

    for i, m in enumerate(movements):
        row = start_row + 1 + i
        data = [
            m.cree_le.strftime('%d/%m/%Y %H:%M'),
            m.produit.nom,
            m.produit.reference,
            str(m.entrepot),
            type_map.get(m.type_mouvement, m.type_mouvement),
            m.quantite,
            str(m.effectue_par) if m.effectue_par else '',
            m.motif or '',
        ]
        for col, v in enumerate(data, 1):
            ws.cell(row=row, column=col, value=v)
        _style_xl_row(ws, row, len(headers), alt=(i % 2 == 1))

    widths = [18, 32, 14, 18, 14, 12, 18, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_requisition_excel(lignes, company=None, language='fr'):
    """Fiche de réquisition provisoire en Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Réquisition'

    start_row = _xl_company_header(ws, company, 'FICHE DE RÉQUISITION PROVISOIRE')

    if not lignes:
        ws.cell(row=start_row, column=1, value='✓ Aucun produit en alerte. Le stock est suffisant.').font = Font(italic=True, color=SUCCESS.lstrip('#').upper())
    else:
        total = sum(l['estimation'] for l in lignes)
        # Bandeau alerte
        ws.cell(row=start_row, column=1, value=f"⚠ {len(lignes)} produit(s) en alerte").font = Font(bold=True, color='FFFFFF', size=11)
        ws.cell(row=start_row, column=1).fill = _xl_total_fill()
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=9)
        ws.cell(row=start_row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
        start_row += 1
        ws.cell(row=start_row, column=1, value=f"Estimation totale : {format_montant(total, company)}").font = Font(bold=True)
        start_row += 2

        headers = ['#', 'Référence', 'Produit', 'Catégorie', 'Stock actuel', 'Seuil',
                   'Qté cible', 'Qté à commander', 'Fournisseur', 'Prix unit.', 'Estimation']
        for col, h in enumerate(headers, 1):
            ws.cell(row=start_row, column=col, value=h)
        _style_xl_header(ws, start_row, len(headers))

        for i, l in enumerate(lignes):
            row = start_row + 1 + i
            data = [
                i + 1,
                l['produit'].reference,
                l['produit'].nom,
                l['produit'].categorie.nom if l['produit'].categorie else '',
                l['stock_actuel'],
                l['produit'].seuil_alerte,
                l['qte_cible'],
                l['qte_a_commander'],
                l['fournisseur'].nom if l['fournisseur'] else '—',
                format_montant(l['prix_unitaire'], company),
                format_montant(l['estimation'], company),
            ]
            for col, v in enumerate(data, 1):
                ws.cell(row=row, column=col, value=v)
            _style_xl_row(ws, row, len(headers), alt=(i % 2 == 1))

        total_row = start_row + 1 + len(lignes)
        for col in range(1, 10):
            cell = ws.cell(row=total_row, column=col)
            cell.fill = _xl_total_fill()
            cell.font = Font(bold=True, color='FFFFFF')
            cell.border = _xl_border()
        ws.cell(row=total_row, column=10, value='TOTAL ESTIMÉ').font = Font(bold=True, color='FFFFFF')
        ws.cell(row=total_row, column=10).alignment = Alignment(horizontal='right')
        ws.cell(row=total_row, column=11, value=format_montant(total, company))
        ws.cell(row=total_row, column=11).font = Font(bold=True, color='FFFFFF')

        widths = [5, 14, 30, 18, 12, 10, 10, 14, 22, 14, 16]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
